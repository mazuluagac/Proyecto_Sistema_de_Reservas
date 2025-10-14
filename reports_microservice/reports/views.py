from rest_framework import viewsets
from .models import Reserva
from .serializers import ReservaSerializer
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
import openpyxl
from reportlab.pdfgen import canvas
from datetime import datetime

# CRUD de reservas
class ReservaViewSet(viewsets.ModelViewSet):
    queryset = Reserva.objects.all()
    serializer_class = ReservaSerializer

# Reporte en Excel

class ReporteExcelView(APIView):
    def get(self, request):
        reservas = Reserva.objects.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservas"

        from openpyxl.styles import Font, Alignment, Border, Side

        headers = [
            "ID", "Usuario ID", "Nombre Usuario", "Fecha Inicio", "Fecha Fin",
            "Descripción", "Estado", "Creado", "Actualizado"
        ]
        ws.append(headers)
        # Estilo encabezado
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        # Bordes finos
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for idx, r in enumerate(reservas, start=2):
            row = [
                r.id,
                r.usuario_id,
                r.nombre_usuario,
                str(r.fecha_inicio),
                str(r.fecha_fin),
                r.descripcion,
                r.estado,
                str(r.created_at),
                str(r.updated_at)
            ]
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left")

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Nombre con timestamp para evitar conflictos
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_reservas_{timestamp}.xlsx"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

# Reporte en PDF

class ReportePDFView(APIView):
    def get(self, request):
        reservas = Reserva.objects.all()
        
        # Nombre con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_reservas_{timestamp}.pdf"

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        p = canvas.Canvas(response, pagesize=(800, 1000))
        width, height = 800, 1000
        p.setFont("Helvetica-Bold", 18)
        p.drawCentredString(width//2, height-50, "Reporte de Reservas")
        p.setFont("Helvetica", 10)
        p.drawString(50, height-70, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        y = height-100
        for r in reservas:
            if y < 120:
                p.showPage()
                p.setFont("Helvetica-Bold", 18)
                p.drawCentredString(width//2, height-50, "Reporte de Reservas")
                p.setFont("Helvetica", 10)
                p.drawString(50, height-70, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                y = height-100
            p.setFont("Helvetica-Bold", 11)
            p.drawString(50, y, f"ID: {r.id} | Usuario ID: {r.usuario_id} | Nombre: {r.nombre_usuario}")
            y -= 16
            p.setFont("Helvetica", 10)
            p.drawString(60, y, f"Fecha inicio: {r.fecha_inicio}  |  Fecha fin: {r.fecha_fin}")
            y -= 14
            p.drawString(60, y, f"Descripción: {r.descripcion}")
            y -= 14
            p.drawString(60, y, f"Estado: {r.estado}")
            y -= 14
            p.drawString(60, y, f"Creado: {r.created_at}  |  Actualizado: {r.updated_at}")
            y -= 14
            # Línea divisoria
            p.setStrokeColorRGB(0.7,0.7,0.7)
            p.line(45, y, width-45, y)
            y -= 18

        p.showPage()
        p.save()
        return response
